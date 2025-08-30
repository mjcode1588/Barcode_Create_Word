from openpyxl import load_workbook, Workbook
from typing import List, Tuple, Optional, Dict
from src.models.product import Product
import os

class ExcelService:
    """Excel 파일 읽기/쓰기 서비스 (CRUD 지원)"""
    
    def __init__(self, file_path: str = "data/items.xlsx"):
        self.file_path = file_path
        self.category_name_to_id: Dict[str, int] = {}
        self.category_id_to_name: Dict[int, str] = {}
        self._ensure_file_exists()
        self._load_categories()  # 시작 시 TYPE 목록 로드
    
    def _ensure_file_exists(self):
        """Excel 파일이 존재하지 않으면 기본 구조로 생성"""
        if not os.path.exists(self.file_path):
            self._create_default_file()
    
    def _create_default_file(self):
        """기본 Excel 파일 생성 (product와 type 시트 포함)"""
        wb = Workbook()
        
        product_ws = wb.active
        product_ws.title = "product"
        product_headers = ["PRODUCT", "PRICE", "TYPE_ID", "PRODUCT_ID"]
        product_ws.append(product_headers)
        
        for col in range(1, len(product_headers) + 1):
            cell = product_ws.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)
        
        type_ws = wb.create_sheet("type")
        type_headers = ["TYPE", "TYPE_ID"]
        type_ws.append(type_headers)
        
        for col in range(1, len(type_headers) + 1):
            cell = type_ws.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)

        default_categories = ["폰스트랩", "리본 키링", "미니 키링", "키링", "팔찌", "꽃갈피", "모양", "부착"]
        for idx, type_name in enumerate(default_categories):
            type_ws.append([type_name, idx])
        
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        wb.save(self.file_path)
        wb.close()
        print(f"기본 Excel 파일 생성: {self.file_path}")
    
    def _load_categories(self):
        """type 시트에서 TYPE 목록 로드"""
        try:
            wb = load_workbook(self.file_path)
            
            if "type" not in wb.sheetnames:
                print("type 시트가 없습니다. 기본 TYPE를 사용합니다.")
                default_categories = ["폰스트랩", "리본 키링", "미니 키링", "키링", "팔찌", "꽃갈피", "모양", "부착"]
                self.category_name_to_id = {name: i for i, name in enumerate(default_categories)}
                self.category_id_to_name = {i: name for i, name in enumerate(default_categories)}
                wb.close()
                return
            
            type_ws = wb["type"]
            
            self.category_name_to_id = {}
            self.category_id_to_name = {}
            for row in type_ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None and row[1] is not None:
                    type_name = str(row[0]).strip()
                    try:
                        type_id = int(row[1])
                        if type_name:
                            self.category_name_to_id[type_name] = type_id
                            self.category_id_to_name[type_id] = type_name
                    except (ValueError, TypeError):
                        continue

            wb.close()
            print(f"TYPE 목록 로드됨: {list(self.category_name_to_id.keys())}")
            
        except Exception as e:
            print(f"TYPE 목록 로드 실패: {e}")
            default_categories = ["폰스트랩", "리본 키링", "미니 키링", "키링", "팔찌", "꽃갈피", "모양", "부착"]
            self.category_name_to_id = {name: i for i, name in enumerate(default_categories)}
            self.category_id_to_name = {i: name for i, name in enumerate(default_categories)}
    
    def get_categories(self) -> Dict[str, int]:
        """TYPE 목록 반환 (이름 -> ID 맵)"""
        return self.category_name_to_id.copy()

    def get_all_categories(self) -> List[str]:
        return list(self.category_name_to_id.keys())

    def is_type_name_in_use(self, type_name: str) -> bool:
        """해당 TYPE가 상품에서 사용 중인지 확인"""
        products = self.read_products()
        return any(p.type_name == type_name for p in products)

    def update_type_name(self, old_type_name: str, new_type_name: str) -> bool:
        """TYPE 수정 (연관된 모든 상품 정보 포함)"""
        if not new_type_name or new_type_name in self.category_name_to_id:
            return False
        
        try:
            wb = load_workbook(self.file_path)
            
            type_ws = wb["type"]
            for row in type_ws.iter_rows(min_row=2):
                if row[0].value == old_type_name:
                    row[0].value = new_type_name
                    break
            
            wb.save(self.file_path)
            wb.close()
            
            self._load_categories() # Reload categories from file
            print(f"TYPE 수정 완료: '{old_type_name}' -> '{new_type_name}'")
            return True
            
        except Exception as e:
            print(f"TYPE 수정 실패: {e}")
            return False

    def delete_type_name(self, type_name: str) -> bool:
        """TYPE 삭제"""
        if self.is_type_name_in_use(type_name):
            print(f"'{type_name}' TYPE는 현재 사용 중이므로 삭제할 수 없습니다.")
            return False
        
        try:
            wb = load_workbook(self.file_path)
            type_ws = wb["type"]
            
            row_to_delete = None
            for row_idx, row in enumerate(type_ws.iter_rows(min_row=2), start=2):
                if row[0].value == type_name:
                    row_to_delete = row_idx
                    break
            
            if row_to_delete:
                type_ws.delete_rows(row_to_delete)
            
            wb.save(self.file_path)
            wb.close()
            
            self._load_categories() # Reload categories
            
            print(f"TYPE 삭제 완료: '{type_name}'")
            return True
            
        except Exception as e:
            print(f"TYPE 삭제 실패: {e}")
            return False

    def add_type_name(self, type_name: str) -> bool:
        """새 TYPE 추가"""
        try:
            if type_name in self.category_name_to_id:
                return True
            
            wb = load_workbook(self.file_path)
            
            if "type" not in wb.sheetnames:
                type_ws = wb.create_sheet("type")
                type_ws.append(["TYPE", "TYPE_ID"])
            else:
                type_ws = wb["type"]
            
            # Find max type_id and add 1
            max_id = -1
            for row in type_ws.iter_rows(min_row=2, values_only=True):
                if row and row[1] is not None:
                    try:
                        max_id = max(max_id, int(row[1]))
                    except (ValueError, TypeError):
                        continue
            new_id = max_id + 1
            
            type_ws.append([type_name, new_id])
            
            wb.save(self.file_path)
            wb.close()
            
            self.category_name_to_id[type_name] = new_id
            self.category_id_to_name[new_id] = type_name
            print(f"새 TYPE 추가됨: {type_name} (ID: {new_id})")
            return True
            
        except Exception as e:
            print(f"TYPE 추가 실패: {e}")
            return False
    
    def read_products(self) -> List[Product]:
        """product 시트에서 상품 정보 읽기 (Read)"""
        try:
            wb = load_workbook(self.file_path, data_only=True)
            
            if "product" in wb.sheetnames:
                ws = wb["product"]
            else:
                ws = wb.active
                print("product 시트가 없어서 첫 번째 시트를 사용합니다.")
            
            products = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                
                try:
                    name = str(row[0]).strip()
                    price = str(row[1]).strip() if row[1] is not None else "0"
                    type_id = int(row[2]) if row[2] is not None else None
                    product_id = int(row[3]) if len(row) > 3 and row[3] is not None else 0
                    
                    if name and type_id is not None:
                        type_name = self.category_id_to_name.get(type_id, "알 수 없음")
                        product = Product(
                            name=name,
                            price=price,
                            type_name=type_name,
                            type_id=type_id,
                            product_id=product_id,
                            barcode_num=f"{type_id}{str(product_id).zfill(6)}"
                        )
                        products.append(product)
                except (ValueError, TypeError, IndexError) as e:
                    print(f"상품 데이터 오류 (행 스킵): {row} - {e}")
                    continue
            
            wb.close()
            print(f"총 {len(products)}개 상품 로드됨")
            return products
            
        except Exception as e:
            print(f"Excel 파일 읽기 실패: {e}")
            return []
    
    def save_products(self, products: List[Product], file_path:str) -> bool:
        """상품 목록을 product 시트에 저장 (Create/Update)"""
        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
            else:
                wb = Workbook()
            
            if "product" in wb.sheetnames:
                wb.remove(wb["product"])

            if "type" in wb.sheetnames:
                wb.remove(wb["type"])

            product_ws = wb.create_sheet("product", 0)
            type_ws = wb.create_sheet("type", 0)

            p_headers = ["PRODUCT", "PRICE", "TYPE_ID", "PRODUCT_ID"]
            t_headers = ["TYPE","TYPE_ID"]
            product_ws.append(p_headers)
            type_ws.append(t_headers)
            

            for col in range(1, len(p_headers) + 1):
                cell = product_ws.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
            for col in range(1, len(t_headers) + 1):
                cell = type_ws.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
            
            # Build TYPE mapping from products while avoiding duplicates/conflicts
            types_map: Dict[int, str] = {}
            name_to_id: Dict[str, int] = {}

            for product in products:
                try:
                    tid = int(getattr(product, "type_id", 0) or 0)
                except Exception:
                    tid = 0
                tname = getattr(product, "type_name", None) or getattr(product, "category", "") or f"TYPE_{tid}"

                # If this id already mapped, keep first occurrence (log on conflict)
                if tid in types_map:
                    if types_map[tid] != tname:
                        print(f"TYPE ID 충돌: {tid} '{types_map[tid]}' vs '{tname}' - 기존 이름 유지")
                    continue

                # If this name already mapped to a different id, keep first occurrence (log)
                if tname in name_to_id:
                    existing_id = name_to_id[tname]
                    if existing_id != tid:
                        print(f"TYPE 이름 충돌: '{tname}' 이미 ID {existing_id}에 할당되어 있음 (시도한 ID: {tid}) - 기존 매핑 유지")
                        continue

                types_map[tid] = tname
                name_to_id[tname] = tid

            # Also include any known categories from self (preserve existing TYPE list)
            for name, tid in self.category_name_to_id.items():
                if tid not in types_map and name not in name_to_id:
                    types_map[tid] = name
                    name_to_id[name] = tid

            # Write types sorted by TYPE_ID for predictability
            for tid, tname in sorted(types_map.items(), key=lambda x: x[0]):
                type_ws.append([tname, tid])

            # Write product rows
            for product in products:
                try:
                    product_ws.append([
                        product.name,
                        product.price,
                        int(getattr(product, "type_id", 0) or 0),
                        int(getattr(product, "product_id", 0) or 0)
                    ])
                except Exception:
                    # fallback: write raw values if conversion fails
                    product_ws.append([
                        getattr(product, "name", ""),
                        getattr(product, "price", ""),
                        getattr(product, "type_id", ""),
                        getattr(product, "product_id", "")
                    ])
            
            # Remove default empty sheet if present
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb["Sheet"])

            wb.save(file_path)
            wb.close()

            # Update in-memory category maps to reflect saved TYPE sheet
            self.category_id_to_name = {int(k): v for k, v in types_map.items()}
            self.category_name_to_id = {v: int(k) for k, v in types_map.items()}

            print(f"Excel 파일 저장 완료: {file_path} ({len(products)}개 상품, {len(types_map)}개 TYPE)")
            return True
            
        except Exception as e:
            print(f"Excel 파일 저장 실패: {e}")
            return False
    
    def add_product(self, product: Product) -> bool:
        """새 상품 추가 (Create)"""
        try:
            products = self.read_products()
            products.append(product)
            return self.save_products(products, self.file_path)
        except Exception as e:
            print(f"상품 추가 실패: {e}")
            return False
    
    def update_product(self, old_product: Product, new_product: Product) -> bool:
        """상품 정보 수정 (Update)"""
        try:
            products = self.read_products()
            
            for i, p in enumerate(products):
                if p.product_id == old_product.product_id:
                    products[i] = new_product
                    break
            
            return self.save_products(products, self.file_path)
        except Exception as e:
            print(f"상품 수정 실패: {e}")
            return False
    
    def delete_product(self, product: Product) -> bool:
        """상품 삭제 (Delete)"""
        try:
            products = self.read_products()
            
            products = [p for p in products if p.product_id != product.product_id or p.type_id != product.type_id or p.name != product.name]
            
            return self.save_products(products, self.file_path)
        except Exception as e:
            print(f"상품 삭제 실패: {e}")
            return False
    
    def get_product_by_name_type_name(self, name: str, type_name: str) -> Optional[Product]:
        # This method might be problematic if type_name is not unique
        # It's better to use ID if possible
        products = self.read_products()
        for product in products:
            if product.name == name and product.type_name == type_name:
                return product
        return None
    
    def get_type_name_counters(self, products: List[Product]) -> dict:
        type_name_counters = {}
        for product in products:
            if product.type_name not in type_name_counters:
                type_name_counters[product.type_name] = 0
            type_name_counters[product.type_name] += 1
        return type_name_counters
    
    def generate_barcode_numbers(self, products: List[Product]) -> List[Tuple[str, str, str, str]]:
        items = []
        for product in products:

            barcode_format = f"PPON-{str(product.barcode_num)}"
            # This logic seems buggy, it adds 78 + 1 items
            items.append((product.name, product.formatted_price, product.type_name, barcode_format))

        print(f"총 {len(items)}개 라벨 생성됨")
        return items
    
    def backup_file(self, backup_path: str = None) -> bool:
        """Excel 파일 백업"""
        try:
            if backup_path is None:
                import datetime
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_dir = os.path.join(os.path.dirname(self.file_path), 'backup')
                os.makedirs(backup_dir, exist_ok=True)
                backup_path = os.path.join(backup_dir, f"{os.path.basename(self.file_path)}_{timestamp}.bak")

            import shutil
            shutil.copy2(self.file_path, backup_path)
            print(f"파일 백업 완료: {backup_path}")
            return True
            
        except Exception as e:
            print(f"파일 백업 실패: {e}")
            return False